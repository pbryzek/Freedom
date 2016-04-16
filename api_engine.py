from get_search_results_api import APIGetSearchResultsRequest
from get_chart_api import APIGetChartRequest
from get_deep_comps_api import APIGetDeepCompsRequest
from get_comps_api import APIGetCompsRequest
from get_deep_search_results_api import APIGetDeepSearchResultsRequest
from sf_bridge import SFBridge
from openpyxl import load_workbook

import consts.paths as paths
import consts.switches as switches
import common.globals as globals
import openpyxl
import shutil
import os
import time

from common.globals import handle_err_msg
from common.globals import format_citystatezip

class APIEngine(object):

    newline = "\n"
    csv_path = ""
    csv_path_tier2 = ""   
 
    def __init__(self, address, city, state, zip, redfin_link, dom, listing_id, type, timestamp, num_hot_words, price):
        self.address = address
        self.state = state
        self.city = city
        self.zip = zip
        self.citystatezip = format_citystatezip(city, state, zip) 
 
        self.redfin_link = redfin_link
        self.dom = dom 
        self.listing_id = listing_id
        self.type = type
        self.num_hot_words = num_hot_words
        self.timestamp = timestamp
        self.price = price

        if type == switches.PROPERTY_TYPE_REDFIN:
            self.csv_path = paths.RESULTS_PATH + paths.REDFIN_RESULTS_PATH + "_" + str(timestamp) + paths.CSV_ENDING
            self.csv_path_tier2 = paths.RESULTS_PATH + paths.REDFIN_RESULTS_PATH_TIER2 + "_" + str(timestamp) + paths.CSV_ENDING
            

    #Based on the MAO, calculate the commision
    def calculate_commision(self, mao_no_commission):
        if (mao_no_commission - switches.COMMISSION_POINT_3) > switches.PRICE_POINT_2:
            return switches.COMMISSION_POINT_3
        elif (mao_no_commission - switches.COMMISSION_POINT_2) > switches.PRICE_POINT_1:
            return switches.COMMISSION_POINT_2
        elif (mao_no_commission - switches.COMMISSION_POINT_1) > switches.PRICE_POINT_0:
            return switches.COMMISSION_POINT_1
        else:
            return switches.COMMISSION_POINT_0

    def format_float_for_excel(self, str_val):
        str_val = str_val.strip()
        if str_val:
            int_val = float(str_val)
            return int_val
        return str_val

    def format_int_for_excel(self, str_val):
        str_val = str_val.strip()
        if str_val:
            int_val = int(str_val)
            return int_val
        return str_val

    def create_csv(self, home, comps):
        file_path = ""
        if home.tier == 1:
            file_path = self.csv_path
        else:
            file_path = self.csv_path_tier2

        with open(file_path, "a") as csvfile: 
            main_csv_string = ""
       
            title_comma_separate = "prop type, address #, address st, city, state, zip, dom, listing_id, beds, baths, sqfootage, lotsize, yearbuilt, latitude, longitude, redfin link, chartlink, homelink, graphlink, maplink, compslink, zpid, zestimate, lastupdated, rentestimate, lastupdated_rent, distance, # hot words, sold price, sold date, comp score, $/sqft"

            main_csv_string += title_comma_separate
            main_csv_string += self.newline
            
            comma_separated = home.create_csv()
            main_csv_string += comma_separated
            main_csv_string += self.newline

            handle_err_msg("Evaluating " + str(len(comps)) + " valid comps")
            principal_arv = 0
            sqft_aggregate_value = 0
           
            sqft_prices = [] 
            for comp in comps:
                comp_comma_separated = comp.create_csv()
                main_csv_string += comp_comma_separated
                main_csv_string += self.newline

                sqft_value = int(comp.sqftprice)
                sqft_prices.append(sqft_value)

            num_comps = len(comps)
            if num_comps == 0:
                return

            sqft_prices.sort()
            num_tops = 3
            top1 = sqft_prices.pop()
            top2 = sqft_prices.pop()
            top3 = sqft_prices.pop()
            top4 = 0
            top5 = 0

            if num_comps > 10:
                top4 = sqft_prices.pop()
                num_tops += 1

            if num_comps > 20:
                top5 = sqft_prices.pop()
                num_tops += 1

            total_tops = top1 + top2 + top3 + top4 + top5

            #Get the average price/sqft for comps sold
            avg_sqft_price = total_tops / num_tops

            principal_arv = 0
            principal_sqfootage = 0
            if home.sqfootage.strip() != '':
                principal_sqfootage = int(home.sqfootage)
                principal_arv = avg_sqft_price * principal_sqfootage
            repair_light = (principal_sqfootage * switches.LIGHT_REHAB)
            repair_med = (principal_sqfootage * switches.MEDIUM_REHAB)
            repair_high = (principal_sqfootage * switches.HIGH_REHAB)

            commision = 0
            if principal_arv < switches.PRICE_POINT_0:
                commision = switches.COMMISSION_POINT_0
            elif principal_arv < switches.PRICE_POINT_1:
                commision = switches.COMMISSION_POINT_1
            elif principal_arv < switches.PRICE_POINT_2:
                commision = switches.COMMISSION_POINT_2
            else:
                commision = switches.COMMISSION_POINT_3

            arv = (switches.ARV_PERCENTAGE * principal_arv)

            mao_light_no_commison = arv - repair_light
            comission_light = self.calculate_commision(mao_light_no_commison)
            mao_light = mao_light_no_commison - comission_light

            mao_med_no_commison = arv - repair_med
            comission_med = self.calculate_commision(mao_med_no_commison)
            mao_med = mao_med_no_commison - comission_med

            mao_high_no_commison = arv - repair_high
            comission_high = self.calculate_commision(mao_high_no_commison)
            mao_high = mao_high_no_commison - comission_high

            #If the mao_med is greater than our max price, skip this home
            if switches.MAX_PRICE < mao_med:
                return

            #This property passed all the tests, so input it into SalesForce
            sf_bridge = SFBridge()
            listing_id = sf_bridge.create_listing_in_sf(home, mao_high, mao_med, mao_light, principal_arv, repair_high, repair_med, repair_light, avg_sqft_price)    
            if listing_id is None:
                handle_err_msg("SF Bridge, listing id was None, returning here") 
                return

            self.save_to_jason_excel(home, comps, comission_med)

            for comp in comps:
                sf_bridge.create_comp_in_sf(comp, listing_id)

            #Finally write out the string
            csvfile.write(main_csv_string)

            mao_header = "principal_arv,avg $/sqft,rehab_light,rehab_med,rehab_high,mao_light,mao_med,mao_high"
            mao_comma_separated = str(principal_arv) + "," + str(avg_sqft_price) + "," + str(repair_light) + "," + str(repair_med) + "," + str(repair_high) + "," + str(mao_light) + "," + str(mao_med) + "," + str(mao_high)

            csvfile.write(mao_header)
            csvfile.write(self.newline)
            csvfile.write(mao_comma_separated)
            csvfile.write(self.newline)
            csvfile.write(self.newline)

            
    def save_to_jason_excel(self, home, comps, comission_med):
        if switches.JASON_ENABLED:
            new_dir_name = paths.JASON_RESULTS_DIR + str(self.timestamp)
            if not os.path.isdir(new_dir_name):
                os.makedirs(new_dir_name)
            jason_excel_sheet = new_dir_name + "/" + self.address + paths.XLSX_ENDING
            #Copy over template to the results dir
            shutil.copy2(paths.JASON_TEMPLATE_PATH, jason_excel_sheet)

            wb = load_workbook(jason_excel_sheet)
            ws = wb.get_sheet_by_name("SFR ANALYSIS")

            full_address = home.address + " " + home.citystatezip
            ws.cell(row=10, column=2).value = full_address
            ws.cell(row=11, column=2).value = self.price

            if self.type == switches.PROPERTY_TYPE_REDFIN:
                ws.cell(row=12, column=2).value = home.listing_id

            ws.cell(row=14, column=2).value = self.format_int_for_excel(home.beds)
            ws.cell(row=15, column=2).value = self.format_float_for_excel(home.baths)

            today_date = time.strftime('%b %d, %Y')
            ws.cell(row=16, column=2).value = today_date

            ws.cell(row=22, column=2).value = self.format_int_for_excel(home.sqfootage)
            ws.cell(row=23, column=2).value = self.format_int_for_excel(home.lotsize)
            ws.cell(row=24, column=2).value = self.format_int_for_excel(home.yearbuilt)
            ws.cell(row=27, column=2).value = str((100 * switches.SQ_FOOTAGE_PERCENTAGE)) + "%"

            row = 39
            col = 1
            for comp in comps:
                if row > 49:
                    break

                full_address = comp.home.address + " " + comp.home.citystatezip
                ws.cell(row=row, column=1).value = full_address
                ws.cell(row=row, column=2).value = self.format_int_for_excel(comp.soldprice)
                ws.cell(row=row, column=3).value = self.format_int_for_excel(comp.home.sqfootage)
                ws.cell(row=row, column=5).value = self.format_int_for_excel(comp.home.lotsize)
                ws.cell(row=row, column=6).value = self.format_int_for_excel(comp.home.yearbuilt)
                ws.cell(row=row, column=7).value = comp.home.dom
                ws.cell(row=row, column=8).value = comp.solddate
                ws.cell(row=row, column=9).value = comp.distance
                ws.cell(row=row, column=11).value = self.type

                row += 1

            ws.cell(row=81, column=2).value = comission_med
            ws.cell(row=93, column=3).value = comission_med
            ws.cell(row=93, column=4).value = switches.ARV_PERCENTAGE

            wb.save(jason_excel_sheet)


    def make_requests(self):
        handle_err_msg("")
        handle_err_msg("Getting info for property: " + self.address + " " + self.citystatezip)
        #First get the zpid of the home in question

        get_deep_search_api = APIGetDeepSearchResultsRequest(self.type, self.address, self.city, self.state, self.zip, self.dom, self.listing_id, self.num_hot_words, switches.RENTZESTIMATE)
        homes = get_deep_search_api.request()
        if len(homes) == 0:
            handle_err_msg("Get Deep Search Results API did not return any results!")
            return

        home = homes[0]
        home.redfin_link = self.redfin_link

        zpid = home.zpid

        #Get the deep comps for that home
        deep_comps_api = APIGetDeepCompsRequest(self.type, zpid, switches.NUM_COMPS_REQUESTED, home.latitude, home.longitude, home.sqfootage)
        deep_comps, tier = deep_comps_api.request()

        num_comps = len(deep_comps)
        if num_comps < switches.MIN_NUM_COMPS:
            handle_err_msg("A total of " + str(num_comps) + " were found, skipping.")
            return 

        #If the home tier has not already been degraded for previous rules, adjust it based on the comp analysis
        if home.tier != 2:
            home.tier = tier
 
        #Finally put the results in a csv 
        self.create_csv(home, deep_comps)
